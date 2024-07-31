# -------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      auto01
#
# Created:     18.09.2022
# Copyright:   (c) auto01 2022
# Licence:     <your licence>
# #------------------------------------------------------------------------------

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath




data_path = join('..', 'nv2024-07-31.xlsx')
data_path = abspath(data_path)


# читаємо excel-файл
wb = load_workbook(filename=data_path, data_only=True, read_only=True)

# читаємо список листів
wsn = list(wb.sheetnames)
print(wsn)


wsdate = None

for i in wsn:
    if wb[i]['I1'].value == "Автор":
        wsdate = i
if wsdate == None:
    raise NotAllData('No date with Автор')

ws = wb[wsdate]

max_rows = ws.max_row
max_cols = ws.max_column
print(max_rows)
print(max_cols)

'''ws.column_dimensions['A'].wigth = 5
        ws.column_dimensions['A'].height = 25

        ws.column_dimensions['B'].wigth = 10
        ws.column_dimensions['C'].wigth = 20
        ws.column_dimensions['D'].wigth = 20
        ws.column_dimensions['E'].wigth = 30
        ws.column_dimensions['F'].wigth = 20
        ws.column_dimensions['G'].wigth = 20
        ws.column_dimensions['H'].wigth = 20
        ws.column_dimensions['I'].wigth = 15
        ws.column_dimensions['J'].wigth = 20
        ws.column_dimensions['K'].wigth = 20
        ws.column_dimensions['L'].wigth = 20
        ws.column_dimensions['M'].wigth = 25
        ws.column_dimensions['N'].wigth = 20
        ws.column_dimensions['O'].wigth = 20
        ws.column_dimensions['P'].wigth = 10
        ws.column_dimensions['Q'].wigth = 40
'''

table_head = [cell.value for cell in next(
    ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=max_cols))]

mandata = {}

for row in ws.iter_rows(min_row=2, min_col=1, max_row=max_rows, max_col=max_cols):
    if len(row) > 0:
        fio = row[8].value
        if fio is not None:
            fiodata = [cell.value for cell in row]
            if fio not in mandata:
                mandata[fio] = []
            mandata[fio].append(fiodata)


for fio in mandata:
    print(f'{len(mandata[fio])} - незакритi ЗН. Автор: {fio} ')

    wb.close()

#  Створення звітів по кожному менеджеру/автору ЗН і Збереження в окрему книгу EXCEL .xlsx

for fio in mandata:
    exname, *_ = fio.split()
    wb = Workbook()
    # получаем активный лист
    ws = wb.active
    ws.title = f'незав.ЗН -{fio}'
    ws.append(table_head)

    for row in mandata[fio]:
        ws.append(row)

    for i in range(1, 29):
        zagl = ws.cell(row=1, column=i)
        zagl.alignment = Alignment(horizontal='center')
       # zagl.fill = PatternFill(fill_type='solid', start_color='5a61f0', end_color='5a61f0')
        zagl.font = Font(bold=True, italic=False, color='000000', size=10)





    '''nmrow = len(mandata[fio])
    for i in range(2, nmrow + 2):
        ws.cell(row=i, column=1).number_format = '##0'
        ws.cell(row=i, column=1).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=2).number_format = 'dd mm yyyy'
        ws.cell(row=i, column=2).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=3).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=4).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

       # ws.cell(row=i, column=5).fill = PatternFill(
      #      fill_type='solid', start_color='000000', end_color='000000')

        ws.cell(row=i, column=6).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

     #   ws.cell(row=i, column=7).fill = PatternFill(
     #       fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=8).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=9).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=10).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=11).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=12).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=13).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=14).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=15).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=16).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=17).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=18).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=19).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=20).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=21).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=22).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=23).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=24).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=25).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=26).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')

        ws.cell(row=i, column=27).fill = PatternFill(
            fill_type='solid', start_color='E9FFE6', end_color='E9FFE6')

        ws.cell(row=i, column=28).fill = PatternFill(
            fill_type='solid', start_color='F5FFC3', end_color='F5FFC3')
            '''

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 2
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 33
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 9
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 8
    ws.column_dimensions['Q'].width = 10
    ws.column_dimensions['R'].width = 6
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 18
    ws.column_dimensions['V'].width = 24
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 17
    ws.column_dimensions['Y'].width = 6
    ws.column_dimensions['Z'].width = 10
    ws.column_dimensions['AA'].width = 5
    ws.column_dimensions["AB"].width = 12

    exfilename = join('.', 'DataNV', (exname + '.xlsx'))
    exfilename = abspath(exfilename)
    print(exfilename)

    wb.save(exfilename)
    wb.close()